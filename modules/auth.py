"""
modules/auth.py — Customer authentication and data retrieval.

Reads directly from the Excel workbook (customers.xlsx) which contains
three sheets: Customer, Transaction, Card.

On startup, all sheets are loaded into memory and indexed by Account_No.
Authentication looks up by account number and returns a unified customer
profile including their accounts, linked cards, and recent transactions.
"""

import openpyxl
from pathlib import Path
from datetime import datetime
from config import CUSTOMERS_XLS_PATH


# ── Internal data store (loaded once at import time) ─────────────────────────

def _load_workbook() -> dict:
    """
    Parse all three sheets and return a dict keyed by customer ID,
    each containing all their accounts, cards, and transactions.
    """
    wb = openpyxl.load_workbook(CUSTOMERS_XLS_PATH)

    # ── Parse Customer sheet ─────────────────────────────────────────────────
    customers_by_id: dict[int, dict] = {}
    accounts_by_no:  dict[int, dict] = {}

    ws_cust = wb["Customer"]
    headers = [c.value for c in next(ws_cust.iter_rows(min_row=1, max_row=1))]

    for row in ws_cust.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        cid    = record["ID"]
        acc_no = record["Account_No"]

        account = {
            "account_no":          str(acc_no),
            "account_name":        record["Account_Name"],
            "currency":            record["Currency"],
            "account_type":        record["Account_Type"],
            "product_type":        record["Product_Type"],
            "product_description": record["Product_Description"],
            "current_balance":     record["Current_Balance"],
            "account_open_date":   record["Account_Open_Date"],
        }
        accounts_by_no[acc_no] = {**account, "customer_id": cid}

        if cid not in customers_by_id:
            customers_by_id[cid] = {
                "id":       cid,
                "name":     record["Account_Name"],
                "accounts": [],
                "cards":    [],
                "transactions_by_account": {},
            }
        customers_by_id[cid]["accounts"].append(account)

    # ── Parse Card sheet ──────────────────────────────────────────────────────
    ws_card = wb["Card"]
    card_headers = [c.value for c in next(ws_card.iter_rows(min_row=1, max_row=1))]

    for row in ws_card.iter_rows(min_row=2, values_only=True):
        rec    = dict(zip(card_headers, row))
        acc_no = rec["Account_No"]

        card = {
            "account_no":         str(acc_no),
            "card_issuer":        rec["Card_Issuer"],
            "card_type":          rec["Card_Type"],
            "card_activation_date": rec["Card_Activation_Date"],
            "status":             rec["Status"],
        }
        # Attach card to the owning customer
        if acc_no in accounts_by_no:
            cid = accounts_by_no[acc_no]["customer_id"]
            customers_by_id[cid]["cards"].append(card)

    # ── Parse Transaction sheet ───────────────────────────────────────────────
    ws_txn = wb["Transaction"]
    txn_headers = [c.value for c in next(ws_txn.iter_rows(min_row=1, max_row=1))]

    for row in ws_txn.iter_rows(min_row=2, values_only=True):
        rec    = dict(zip(txn_headers, row))
        acc_no = rec["Account_No"]

        txn = {
            "date":            rec["Transaction_Date"],
            "type":            rec["Transaction_Type"],
            "amount":          rec["Transaction_Amount"],
            "destination":     str(rec["Destination_Account"]),
            "narration":       rec["Narration"],
            "dest_bank":       rec["Destination_Account_Bank"],
            "status":          rec["Transaction_Status"],
            "failure_reason":  rec["Failure_Reason"],
        }
        if acc_no in accounts_by_no:
            cid = accounts_by_no[acc_no]["customer_id"]
            txns = customers_by_id[cid]["transactions_by_account"]
            txns.setdefault(str(acc_no), []).append(txn)

    return customers_by_id, accounts_by_no


_CUSTOMERS_BY_ID, _ACCOUNTS_BY_NO = _load_workbook()


# ── Public interface ──────────────────────────────────────────────────────────

def get_customer(account_number: str) -> dict | None:
    """
    Look up a customer by account number.

    Args:
        account_number: The account number entered by the user.

    Returns:
        A unified customer profile dict, or None if not found.
    """
    try:
        acc_no = int(account_number.strip())
    except ValueError:
        return None

    if acc_no not in _ACCOUNTS_BY_NO:
        return None

    cid = _ACCOUNTS_BY_NO[acc_no]["customer_id"]
    return _CUSTOMERS_BY_ID.get(cid)


def format_customer_context(customer: dict, queried_account: str) -> str:
    """
    Build a structured text block describing the customer's full profile
    for injection into the LLM system prompt.

    Includes: all accounts, cards, and last 5 transactions for each account.
    """
    lines = [f"Customer Name: {customer['name']}"]

    # ── Accounts ──────────────────────────────────────────────────────────────
    lines.append("\nACCOUNTS:")
    for acc in customer["accounts"]:
        opened = acc["account_open_date"]
        if isinstance(opened, datetime):
            opened = opened.strftime("%Y-%m-%d")
        balance = acc["current_balance"]
        currency = acc["currency"]
        lines.append(
            f"  • Account No: {acc['account_no']} | {acc['product_description']} "
            f"({acc['account_type']}, {currency}) | "
            f"Balance: {currency} {balance:,.2f} | Opened: {opened}"
        )

    # ── Cards ──────────────────────────────────────────────────────────────────
    if customer["cards"]:
        lines.append("\nCARDS:")
        for card in customer["cards"]:
            activated = card["card_activation_date"]
            if isinstance(activated, datetime):
                activated = activated.strftime("%Y-%m-%d")
            lines.append(
                f"  • Account {card['account_no']} | {card['card_issuer']} "
                f"{card['card_type']} Card | Status: {card['status']} | "
                f"Activated: {activated}"
            )

    # ── Recent transactions ────────────────────────────────────────────────────
    txns_map = customer.get("transactions_by_account", {})
    if txns_map:
        lines.append("\nRECENT TRANSACTIONS (last 5 per account):")
        for acc_no, txns in txns_map.items():
            # Sort by date descending, take last 5
            sorted_txns = sorted(
                txns,
                key=lambda t: t["date"] if isinstance(t["date"], datetime) else datetime.min,
                reverse=True
            )[:5]

            lines.append(f"  Account {acc_no}:")
            for t in sorted_txns:
                dt = t["date"].strftime("%Y-%m-%d %H:%M") if isinstance(t["date"], datetime) else str(t["date"])
                amount = abs(t["amount"])
                sign   = "+" if t["type"] == "Credit" else "-"
                reason = f" | Failed: {t['failure_reason']}" if t["failure_reason"] else ""
                lines.append(
                    f"    [{dt}] {t['type']} {sign}{amount:,.2f} — "
                    f"{t['narration']} ({t['dest_bank']}) [{t['status']}{reason}]"
                )

    return "\n".join(lines)
